#!/usr/bin/env python3
"""直接从Excel解析数据并通过REST API批量导入MemFire数据库"""
import openpyxl
import json
import urllib.request
import urllib.error
import sys
from uuid import uuid4

EXCEL_PATH = '/Users/jyf/.hermes/cache/documents/doc_b48cd57bd32a_Yangtze - 事项进度-甘特视图.xlsx'
ENV_PATH = '/Users/jyf/Desktop/ipo-management-system/.env.local'

env = {}
with open(ENV_PATH, 'r') as f:
    for line in f:
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            key, val = line.split('=', 1)
            env[key] = val

SUPABASE_URL = env['NEXT_PUBLIC_SUPABASE_URL']
SUPABASE_KEY = env['NEXT_PUBLIC_SUPABASE_ANON_KEY']

def api(endpoint, method='GET', data=None):
    url = f"{SUPABASE_URL}/rest/v1/{endpoint}"
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=minimal'
    }
    body = json.dumps(data).encode('utf-8') if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return resp.read().decode('utf-8')
    except urllib.error.HTTPError as e:
        print(f"ERROR {e.code}: {e.read().decode()}", file=sys.stderr)
        raise

# 1. 清空
print("清空现有数据...", flush=True)
for table in ['gantt_cells', 'tasks', 'workstreams', 'projects']:
    try:
        api(f'{table}?id=neq.00000000-0000-0000-0000-000000000000', method='DELETE')
    except:
        pass

# 2. 创建项目
project_id = str(uuid4())
api('projects', method='POST', data={'id': project_id, 'name': 'Project Yangtze', 'description': '港股IPO项目'})
print(f"项目已创建: {project_id}", flush=True)

# 3. 解析Excel
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active
header_row = 4  # 已知表头在第4行

headers = [c.value or '' for c in ws[header_row]]

def find_col(name):
    for i, h in enumerate(headers):
        if name in str(h):
            return i
    return None

col_title = find_col('事项')       # E列=4
col_sponsor = find_col('保荐人')   # D列=3
col_lawyer = find_col('律师')      # C列=2
col_other = find_col('其他机构')   # B列=1
col_progress = find_col('当前进度') # F列=5
col_blocker = find_col('当前卡点')  # G列=6
col_next = find_col('下一步')      # H列=7

# 日期列
date_cols = []
for i, h in enumerate(headers):
    if h and '2026' in str(h) or (h and '2025' in str(h)):
        date_str = str(h)[:10]  # 取 YYYY-MM-DD
        date_cols.append((i, date_str))

print(f"列: title={col_title}, sponsor={col_sponsor}, lawyer={col_lawyer}, other={col_other}")
print(f"日期列: {len(date_cols)} 个", flush=True)

# 4. 解析
workstreams = {}
all_tasks = []
all_gantt = []
ws_order = 0
task_order = 0
current_ws_id = None

for row_idx in range(header_row + 1, ws.max_row + 1):
    title_cell = ws.cell(row_idx, col_title + 1)
    title_val = title_cell.value
    if not title_val or not str(title_val).strip():
        continue
    
    title_str = str(title_val).strip()
    is_bold = title_cell.font and title_cell.font.bold
    fill_rgb = ''
    if title_cell.fill and title_cell.fill.start_color:
        fill_rgb = str(title_cell.fill.start_color.rgb or '').upper()
    
    if is_bold and 'FF0C306E' in fill_rgb:
        ws_id = str(uuid4())
        workstreams[title_str] = {'id': ws_id, 'project_id': project_id, 'name': title_str, 'sort_order': ws_order}
        ws_order += 1
        current_ws_id = ws_id
        continue
    
    if not current_ws_id:
        continue
    
    row = list(ws[row_idx])
    task_id = str(uuid4())
    
    def cell_val(col_idx):
        if col_idx is None: return ''
        v = row[col_idx].value
        return str(v).strip() if v else ''
    
    all_tasks.append({
        'id': task_id,
        'project_id': project_id,
        'workstream_id': current_ws_id,
        'title': title_str,
        'sponsor': cell_val(col_sponsor),
        'lawyer': cell_val(col_lawyer),
        'other_party': cell_val(col_other),
        'current_progress': cell_val(col_progress),
        'current_blocker': cell_val(col_blocker),
        'next_step': cell_val(col_next),
        'status': 'in_progress',
        'sort_order': task_order,
        'remark': None,
        'assignee': None
    })
    
    for col_idx, date_str in date_cols:
        cell = row[col_idx]
        fill_rgb = ''
        if cell.fill and cell.fill.start_color:
            fill_rgb = str(cell.fill.start_color.rgb or '').upper()
        
        cell_type = None
        if 'FF1450B8' in fill_rgb:
            cell_type = 'ddl'  # 深蓝=节点/里程碑
        elif 'FF99BEFF' in fill_rgb:
            cell_type = 'event'  # 浅蓝=条线范围
        
        if cell_type:
            label = str(cell.value).strip() if cell.value else ''
            all_gantt.append({
                'id': str(uuid4()),
                'task_id': task_id,
                'cell_date': date_str,
                'label': label,
                'cell_type': cell_type
            })
    
    task_order += 1

print(f"解析完成: {len(workstreams)} 条线, {len(all_tasks)} 任务, {len(all_gantt)} 甘特格", flush=True)

# 5. 批量插入
ws_list = list(workstreams.values())
if ws_list:
    api('workstreams', method='POST', data=ws_list)
    print(f"✅ 条线已插入", flush=True)

# 批量插入任务（每批50个）
BATCH = 50
for i in range(0, len(all_tasks), BATCH):
    batch = all_tasks[i:i+BATCH]
    api('tasks', method='POST', data=batch)
    print(f"  任务 {i+1}-{min(i+BATCH, len(all_tasks))}/{len(all_tasks)}", flush=True)

print(f"✅ 任务已插入", flush=True)

# 批量插入甘特格（每批200个）
GBATCH = 200
for i in range(0, len(all_gantt), GBATCH):
    batch = all_gantt[i:i+GBATCH]
    api('gantt_cells', method='POST', data=batch)
    print(f"  甘特 {i+1}-{min(i+GBATCH, len(all_gantt))}/{len(all_gantt)}", flush=True)

print(f"✅ 全部导入完成！", flush=True)
